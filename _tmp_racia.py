import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\Denis\Desktop\Система\Хранилище\raw\01_Маркетплейс\Конвеер новинок\Товары из Китая 13.05.xlsx",
    read_only=True, data_only=True
)

print("Все листы:", wb.sheetnames)
print()

if "рация" in wb.sheetnames:
    ws = wb["рация"]
    print("ЛИСТ: рация")
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            clean = [str(c)[:80] if c is not None else "" for c in row[:18]]
            print(" | ".join(c for c in clean if c))
else:
    print("Лист 'рация' не найден в локальном xlsx")
    # Search for similar
    for name in wb.sheetnames:
        if "рац" in name.lower() or "radio" in name.lower() or "walk" in name.lower():
            print("Похожий лист:", name)
