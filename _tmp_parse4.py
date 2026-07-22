import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\Denis\Desktop\Система\Хранилище\raw\01_Маркетплейс\Конвеер новинок\Товары из Китая 13.05.xlsx",
    read_only=True, data_only=True
)

# Sheets 18-26 (not seen in GDrive content)
target_sheets = wb.sheetnames[18:]
print("Листы 18+:", target_sheets)
print()

for sh_name in target_sheets:
    ws = wb[sh_name]
    print("=" * 60)
    print("ЛИСТ:", sh_name)
    rows_printed = 0
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            clean = [str(c)[:60] if c is not None else "" for c in row[:18]]
            print(" | ".join(c for c in clean if c))
            rows_printed += 1
            if rows_printed >= 6:
                break
    print()
