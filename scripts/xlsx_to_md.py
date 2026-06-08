"""
Конвертирует xlsx в markdown + извлекает все изображения.
Использование: python xlsx_to_md.py <путь_к_файлу.xlsx>
"""
import sys
import os
import zipfile
import shutil
from pathlib import Path
import openpyxl


def extract_images(xlsx_path: Path, out_dir: Path) -> list[str]:
    """Извлекает изображения из xlsx (они хранятся внутри zip)."""
    images = []
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for name in z.namelist():
            if name.startswith('xl/media/'):
                filename = Path(name).name
                dest = out_dir / filename
                with z.open(name) as src, open(dest, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                images.append(filename)
    return images


def sheet_to_md(ws) -> str:
    """Конвертирует лист в markdown-таблицу."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # Убираем полностью пустые строки с конца
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    if not rows:
        return ""

    # Определяем ширину
    col_count = max(len(r) for r in rows)

    lines = []
    for i, row in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in row]
        # Дополняем до col_count
        cells += [""] * (col_count - len(cells))
        lines.append("| " + " | ".join(cells) + " |")
        if i == 0:
            lines.append("| " + " | ".join(["---"] * col_count) + " |")

    return "\n".join(lines)


def convert(xlsx_path: str):
    xlsx_path = Path(xlsx_path).resolve()
    stem = xlsx_path.stem  # имя без расширения
    out_dir = xlsx_path.parent / stem
    out_dir.mkdir(exist_ok=True)

    # Извлекаем изображения
    images = extract_images(xlsx_path, out_dir)

    # Читаем данные
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    md_lines = [f"# {stem}\n"]
    md_lines.append(f"Источник: `{xlsx_path.name}`\n")

    # Изображения
    if images:
        md_lines.append("## Изображения\n")
        for img in sorted(images):
            md_lines.append(f"![{img}]({img})")
        md_lines.append("")

    # Листы
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"## Лист: {sheet_name}\n")
        table = sheet_to_md(ws)
        if table:
            md_lines.append(table)
        else:
            md_lines.append("_(пусто)_")
        md_lines.append("")

    md_path = out_dir / f"{stem}.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")

    print(f"Готово: {md_path}")
    print(f"Изображений: {len(images)}")
    for img in sorted(images):
        print(f"  - {img}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python xlsx_to_md.py <файл.xlsx>")
        sys.exit(1)
    convert(sys.argv[1])
