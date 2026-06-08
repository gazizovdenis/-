"""
Конвертирует еженедельные отчёты (WB/Ozon xlsx) в единый CSV.
Берёт последний файл для каждого ИП+платформа (в нём полная история).
Читает лист "Сводная" с финансовыми данными.
"""
import re
import csv
import warnings
from pathlib import Path
import openpyxl

warnings.filterwarnings("ignore")

BASE_DIR = Path(r"c:\Users\Denis\Desktop\Система")
REPORTS_DIR = BASE_DIR / "Хранилище/raw/01_Маркетплейс/Еженедельные отчеты"
OUT_CSV = BASE_DIR / "Хранилище/raw/01_Маркетплейс/Еженедельные отчеты/all_reports.csv"

# Колонки для Ozon (0-based index от начала строки)
OZON_COLS = {
    "period":    0,   # Имя периода
    "date":      1,   # Дата
    "week":      2,   # Нед
    "rev_promo": 3,   # Выручка с продвижением
    "u_ordered": 4,   # Шт. заказы
    "u_sold":    5,   # Шт. продажи
    "u_returns": 6,   # Шт. возвраты
    "orders_rub":7,   # По электронному чеку: заказы
    "returns_rub":8,  # По электронному чеку: возвраты
    "net_rub":   9,   # По электронному чеку: итого нетто
    "cost":      10,  # Себестоимость
    "commission":11,  # Комиссия
    "logistics": 12,  # Логистика
}

# Колонки для WB (смещение +2 от Ozon, т.к. А=None, E=None)
WB_COLS = {k: v + 2 if v >= 3 else (v + 1 if v >= 1 else v)
           for k, v in OZON_COLS.items()}
# Корректируем вручную (A=None добавляет +1, E=None ещё +1 для финансовых колонок)
WB_COLS = {
    "period":    1,
    "date":      2,
    "week":      3,
    "rev_promo": 5,
    "u_ordered": 6,
    "u_sold":    7,
    "u_returns": 8,
    "orders_rub":9,
    "returns_rub":10,
    "net_rub":   11,
    "cost":      13,  # Себестоимость WB (сдвиг из-за Средний ВП в col 12)
    "commission":14,  # Комиссия WB
    "logistics": 15,  # Логистика WB
}

CSV_HEADERS = [
    "ип", "платформа", "период", "дата", "неделя",
    "выручка_с_продвижением", "шт_заказы", "шт_продажи", "шт_возвраты",
    "заказы_руб", "возвраты_руб", "нетто_руб",
    "себестоимость", "комиссия", "логистика",
]


def parse_file_sort_key(filename: str) -> tuple:
    """
    Сортировочный ключ: (год, месяц_окончания, неделя).
    Формат в имени: 'NN неделя DD-DD.MM' или 'NN-NN недели DD.MM-DD.MM'
    Текущий год 2026; если месяц > 5, это предыдущий год (2025).
    """
    # Извлекаем конечную дату вида DD.MM или DD.MM.YYYY
    m = re.search(r"(\d{1,2})\.(\d{1,2})(?:\.(\d{4}))?(?:\s|\.xlsx)", filename)
    if m:
        day, month, year_str = m.group(1), int(m.group(2)), m.group(3)
        if year_str:
            year = int(year_str)
        elif month > 5:   # июнь-декабрь → прошлый год
            year = 2025
        else:
            year = 2026
    else:
        month, year = 0, 2024

    # Номер недели
    wm = re.search(r"(\d+)[-–]?\d*\s*н", filename)
    week = int(wm.group(1)) if wm else 0
    # Высокий номер недели (>30) в прошлом году ставим в начало
    if week > 30 and year < 2026:
        week -= 100

    return (year, month, week)


def find_detail_sheet(wb: openpyxl.Workbook, platform: str):
    """Находит лист Сводная с финансовыми данными (>20 колонок)."""
    candidates = [s for s in wb.sheetnames if "Сводная" in s and "план" not in s]
    for name in candidates:
        ws = wb[name]
        if ws.max_column > 20:
            return ws
    # fallback: второй Сводная по порядку
    found = [wb[s] for s in wb.sheetnames if "Сводная" in s]
    return found[-1] if found else None


def extract_rows(ws, cols: dict, ip: str, platform: str) -> list:
    """Извлекает строки данных из листа."""
    rows_out = []
    data_started = False

    for row in ws.iter_rows(values_only=True):
        date_val = row[cols["date"]] if len(row) > cols["date"] else None
        rev_val  = row[cols["rev_promo"]] if len(row) > cols["rev_promo"] else None

        # Пропускаем заголовки и пустые строки
        if not isinstance(date_val, (str, int, float)) or date_val is None:
            continue
        # Пропускаем строку "Итого"
        period_val = row[cols["period"]] if len(row) > cols["period"] else None
        if isinstance(period_val, str) and "итого" in period_val.lower():
            continue
        # Пропускаем нулевые (будущие) недели
        if rev_val == 0 or rev_val is None:
            continue

        def get(key):
            idx = cols[key]
            return row[idx] if len(row) > idx else None

        rows_out.append({
            "ип":                   ip,
            "платформа":            platform,
            "период":               str(get("period") or "").strip(),
            "дата":                 str(date_val).strip(),
            "неделя":               get("week"),
            "выручка_с_продвижением": get("rev_promo"),
            "шт_заказы":            get("u_ordered"),
            "шт_продажи":           get("u_sold"),
            "шт_возвраты":          get("u_returns"),
            "заказы_руб":           get("orders_rub"),
            "возвраты_руб":         get("returns_rub"),
            "нетто_руб":            get("net_rub"),
            "себестоимость":        get("cost"),
            "комиссия":             get("commission"),
            "логистика":            get("logistics"),
        })

    return rows_out


def main():
    # Группируем файлы по (ИП, платформа), берём последний по неделе
    file_groups: dict[tuple, list] = {}
    for xlsx in REPORTS_DIR.rglob("*.xlsx"):
        name = xlsx.name
        m = re.search(r"ИП\s+(\w+)\s+(WB|Ozon)", name, re.IGNORECASE)
        if not m:
            continue
        ip, platform = m.group(1), m.group(2).upper()
        key = (ip, platform)
        file_groups.setdefault(key, []).append(xlsx)

    all_rows = []
    for (ip, platform), files in sorted(file_groups.items()):
        # Берём файл с наибольшей датой окончания периода
        latest = max(files, key=lambda f: parse_file_sort_key(f.name))
        print(f"  {ip} {platform}: {latest.name}")

        wb = openpyxl.load_workbook(latest, data_only=True)
        ws = find_detail_sheet(wb, platform)
        if ws is None:
            print(f"    [!] Не найден лист Сводная")
            continue

        cols = WB_COLS if platform == "WB" else OZON_COLS
        rows = extract_rows(ws, cols, ip, platform)
        print(f"    ok {len(rows)} строк")
        all_rows.extend(rows)

    # Пишем CSV (UTF-8 BOM для Excel)
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\nГотово: {OUT_CSV}")
    print(f"Итого строк: {len(all_rows)}")


if __name__ == "__main__":
    main()
