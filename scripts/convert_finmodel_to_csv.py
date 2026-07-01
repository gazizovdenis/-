"""
Извлекает ключевые метрики из 2026 Финансовая модель.xlsx
Лист "В PL" (индекс 4) — P&L по WB и Ozon помесячно.
Результат: fin_model_pl.csv — строки=метрики, столбцы=месяцы.
"""
import csv
import warnings
from pathlib import Path
import openpyxl

warnings.filterwarnings("ignore")

BASE_DIR = Path(r"c:\Users\Denis\Desktop\Система")
FIN_FILE = BASE_DIR / "Хранилище/raw/01_Маркетплейс/Еженедельные отчеты/2026 Финансовая модель (1).xlsx"
OUT_CSV  = BASE_DIR / "Хранилище/raw/01_Маркетплейс/Еженедельные отчеты/fin_model_pl.csv"

# Индексы столбцов (0-based) для месяцев
MONTHS = {
    "дек":  2,  # 29-31.12
    "янв":  3,  # Январь
    "фев":  4,  # Февраль
    "мар":  5,  # Март
    "апр":  6,  # Апрель
    "май":  7,  # Май
    "июн":  8,
    "июл":  9,
    "авг":  10,
    "сен":  11,
    "окт":  12,
    "ноя":  13,
    "дек2": 14,
}

# Ключевые строки для извлечения: (row_num_1based, метка, категория)
KEY_ROWS = [
    # WB
    (13,  "wb_заказы_шт",                   "WB"),
    (14,  "wb_продажи_шт",                  "WB"),
    (17,  "wb_нетто_руб",                   "WB"),
    (19,  "wb_выручка_всего_руб",            "WB"),
    (21,  "wb_выручка_вп_руб",              "WB"),
    (22,  "wb_маржа_%",                      "WB"),
    (24,  "wb_выручка_после_вычетов_руб",    "WB"),
    (26,  "wb_реклама_вп_руб",               "WB"),
    (34,  "wb_валовая_прибыль_руб",          "WB"),
    (37,  "wb_себестоимость_руб",            "WB"),
    (38,  "wb_себестоимость_%",              "WB"),
    (42,  "wb_реклама_итого_руб",            "WB"),
    (43,  "wb_реклама_%",                    "WB"),
    (44,  "wb_логистика_руб",                "WB"),
    (45,  "wb_реклама_поиск_руб",            "WB"),
    (46,  "wb_реклама_акции_руб",            "WB"),
    (48,  "wb_штрафы_руб",                   "WB"),
    (54,  "wb_расходы_итого_руб",            "WB"),
    # Ozon
    (60,  "ozon_выручка_всего_руб",          "Ozon"),
    (62,  "ozon_нетто_руб",                  "Ozon"),
    (63,  "ozon_нетто_%",                    "Ozon"),
    (65,  "ozon_выручка_после_вычетов_руб",  "Ozon"),
    (67,  "ozon_возвраты_руб",               "Ozon"),
    (68,  "ozon_возвраты_шт",                "Ozon"),
    (71,  "ozon_выручка_без_возвратов_руб",  "Ozon"),
    (77,  "ozon_себестоимость_руб",          "Ozon"),
    (78,  "ozon_себестоимость_%",            "Ozon"),
    (82,  "ozon_реклама_итого_руб",          "Ozon"),
    (83,  "ozon_реклама_%",                  "Ozon"),
    (84,  "ozon_логистика_руб",              "Ozon"),
    (85,  "ozon_логистика_по_единице_руб",   "Ozon"),
    (90,  "ozon_расходы_итого_руб",          "Ozon"),
    # Общие расходы
    (92,  "реклама_wb_ozon_итого_руб",       "Общее"),
    (99,  "зп_логистика_руб",                "Общее"),
    (100, "зп_склад_руб",                    "Общее"),
    (101, "зп_операционные_руб",             "Общее"),
    (102, "mp_stats_руб",                    "Общее"),
    (103, "реклама_внешняя_руб",             "Общее"),
    (104, "реклама_другое_руб",              "Общее"),
    (109, "кредит_ежемесячный_руб",          "Общее"),
    (113, "расходы_операционные_итого_руб",  "Итого"),
    (115, "налоги_руб",                      "Итого"),
    (128, "итого_до_кредита_руб",            "Итого"),
    (129, "кредит_выплата_руб",              "Итого"),
    (130, "чистая_прибыль_руб",              "Итого"),
    (132, "накопленный_cashflow_руб",        "Итого"),
]


def clean_val(v):
    """Приводим значение к числу или пустой строке."""
    if v is None or v == '#DIV/0!' or v == 0:
        return ""
    if isinstance(v, float) and abs(v) < 1e-6:
        return ""
    return round(v, 2) if isinstance(v, float) else v


def main():
    wb = openpyxl.load_workbook(FIN_FILE, data_only=True)
    ws = wb[wb.sheetnames[4]]  # В PL

    # Загружаем все строки в память для быстрого доступа по номеру
    all_rows = list(ws.iter_rows(values_only=True))

    month_keys = list(MONTHS.keys())
    headers = ["метрика", "категория"] + month_keys

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row_num, label, category in KEY_ROWS:
            row = all_rows[row_num - 1]  # 0-based index
            values = [clean_val(row[MONTHS[m]]) for m in month_keys]
            writer.writerow([label, category] + values)

    print(f"Готово: {OUT_CSV}")
    print(f"Строк: {len(KEY_ROWS)}, Месяцев: {len(MONTHS)}")
    print()

    # Preview ключевых итоговых строк
    print("PREVIEW (итоговые строки):")
    with open(OUT_CSV, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        preview_keys = {"чистая_прибыль_руб", "wb_выручка_вп_руб", "ozon_нетто_руб", "расходы_операционные_итого_руб"}
        for row in reader:
            if row["метрика"] in preview_keys:
                vals = [f"{row[m]:>15}" for m in month_keys[:6]]
                print(f"  {row['метрика']:40} | {'|'.join(vals)}")


if __name__ == "__main__":
    main()
