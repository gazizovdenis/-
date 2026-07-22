import sys, io, zipfile, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

path = r"c:\Users\Denis\Desktop\Система\Хранилище\raw\01_Маркетплейс\Конвеер новинок\Товары из Китая 29.06.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("Листы:", wb.sheetnames)
print()

# Known old sheets to skip
old_sheets = {
    'видеокамеры','вибраторы','ночник','Игрушки интерактивные',
    'машинки для стрижки животных','блендеры','детский пылесос','тонометр',
    'конструктор','мультитул','антенны','Шуруповерт','наушники строительные',
    'зонты','бандаж','лежанка','машинка от катышек','отпариватель','павербанк',
    'чехлы для ноута','полка для ванной','стабилизатор','3д ручка','подставка для ноута',
    'Держатели для украшений','тренажер','зарядка','рация','фотоаппарат',
    'швабры','Разделочные доски','ершик'
}

new_sheets = [s for s in wb.sheetnames if s not in old_sheets]
print("НОВЫЕ листы:", new_sheets)
print()

for sh_name in new_sheets:
    ws = wb[sh_name]
    print(f"{'='*60}")
    print(f"ЛИСТ: {sh_name}")
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            clean = [str(c)[:120] if c is not None else "" for c in row[:20]]
            non_empty = [c for c in clean if c.strip()]
            if non_empty:
                print(" | ".join(non_empty))
    print()
