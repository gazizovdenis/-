import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\Denis\Desktop\Система\Хранилище\raw\01_Маркетплейс\Конвеер новинок\Товары из Китая 13.05 (1).xlsx",
    read_only=True, data_only=True
)

print("Листы:", wb.sheetnames)
print()

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    print(f"=== {sh_name} ===")
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            clean = [str(c)[:100] if c is not None else "" for c in row[:20]]
            non_empty = [c for c in clean if c.strip()]
            if non_empty:
                print(" | ".join(non_empty))
    print()
